import os
import pandas as pd
from PIL import Image
from openpyxl import load_workbook
import shutil

def generate_image_metadata():
    metadata = []

    # Read the existing inventory data
    filename = 'book_inventory.xlsx'
    df = pd.read_excel(filename, sheet_name='Sheet1')
    
    processed_images_dir = 'Processed Images'
    os.makedirs(processed_images_dir, exist_ok=True)
    
    for index, row in df.iterrows():
        set_name = row['Title']
        storage = row['Storage'] if 'Storage' in row else 'Uncategorized'
        storage_dir = os.path.join(processed_images_dir, storage)
        os.makedirs(storage_dir, exist_ok=True)

        title_dir = os.path.join(storage_dir, set_name)
        os.makedirs(title_dir, exist_ok=True)

        image_folder = 'Images to be Processed'  # Update this path to the folder where images are initially stored
        image_files = [f for f in os.listdir(image_folder) if f.startswith(set_name)]
        
        for i, image_file in enumerate(sorted(image_files), 1):
            image_path = os.path.join(image_folder, image_file)
            new_image_path = os.path.join(title_dir, image_file)
            shutil.move(image_path, new_image_path)  # Move image to processed folder

            with Image.open(new_image_path) as img:
                metadata.append({
                    "Data Source": "Data by Sar",  # Added field for the data source
                    "Set Name": set_name,
                    "Image Name": image_file,
                    "Format": img.format,
                    "Mode": img.mode,
                    "Size": img.size,
                    "Width": img.width,
                    "Height": img.height,
                    "Position in Set": f"{i} of {len(image_files)}",
                    "Author": row["Author"],
                    "Publisher": row["Publisher"],
                    "Year": row["Year"],
                    "Edition": row["Edition"],
                    "Pages": row["Pages"],
                    "Weight": row["Weight"],
                    "Dimensions": row["Dimensions"],
                    "Format": row["Format"],
                    "ISBN-10": row["ISBN-10"],
                    "ISBN-13": row["ISBN-13"],
                    "Genre": row["Genre"],
                    "Cover Condition": row["Cover Condition"],
                    "Spine Condition": row["Spine Condition"],
                    "Pages Condition": row["Pages Condition"],
                    "Extent of Damage/Markings": row["Extent of Damage/Markings"],
                    "Binding Integrity": row["Binding Integrity"],
                    "Dust Jacket Condition": row["Dust Jacket Condition"],
                    "Markings and Annotations": row["Markings and Annotations"],
                    "Stains and Odours": row["Stains and Odours"],
                    "Total Score": row["Total Score"],
                    "Additional Information": row["Additional Information"]
                })

    df_metadata = pd.DataFrame(metadata)
    
    # Load the existing workbook
    book = load_workbook(filename)
    
    # Check if 'Sheet3' exists, create if it does not
    if 'Sheet3' in book.sheetnames:
        del book['Sheet3']
    sheet3 = book.create_sheet('Sheet3')

    # Write the dataframe to Sheet3
    for r in dataframe_to_rows(df_metadata, index=False, header=True):
        sheet3.append(r)

    # Save the updated workbook
    book.save(filename)
    print(f"Image metadata saved to {filename} in Sheet3")

if __name__ == "__main__":
    generate_image_metadata()
